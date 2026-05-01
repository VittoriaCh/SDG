import os
from pathlib import Path

import numpy as np
import pandas as pd
import rasterio


# ============================================================
# 1. 路径与参数配置
# ============================================================

# 土地覆盖文件所在文件夹
LAND_COVER_DIR = Path(r"C:\deepOptica\deep\fromFeishu\Land_Cover_Each_Year")

# 文件名前缀，例如：
# ESRI_2018_reclass_300m.tif
# ESRI_2019_reclass_300m.tif
DATASET_PREFIX = "C3S"

# 基准年份
BASELINE_YEAR = 2018

# 目标年份
REPORT_YEARS = range(2019, 2025)

# 输出路径
OUTPUT_DIR = Path("./soc_output")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


# ============================================================
# 2. 文件 Table 3.6 中的气候区 f 参数
# ============================================================
# 文件给出的 cropland 相关 f 值：
# Temperate Dry      = 0.80
# Temperate Moist    = 0.69
# Tropical Dry       = 0.58
# Tropical Moist     = 0.48
# Tropical Montane   = 0.64

CLIMATE_F_VALUES = {
    "Temperate Dry": 0.80,
    "Temperate Moist": 0.69,
    "Tropical Dry": 0.58,
    "Tropical Moist": 0.48,
    "Tropical Montane": 0.64,
}

# UAE / Abu Dhabi 通常按 Tropical Dry 使用 f=0.58
CLIMATE_ZONE = "Tropical Dry"
F_VALUE = CLIMATE_F_VALUES[CLIMATE_ZONE]


# ============================================================
# 3. 类别编码
# ============================================================
# 与文件 Table 3.6 对齐：
# 1 = Forest / Tree-covered areas
# 2 = Grasslands
# 3 = Croplands
# 4 = Wetlands
# 5 = Artificial areas
# 6 = Other / Bare lands
# 7 = Water bodies

ALL_CLASSES = [1, 2, 3, 4, 5, 6, 7]
LAND_CLASSES_FOR_AREA = [1, 2, 3, 4, 5, 6]

CLASS_NAMES = {
    1: "Forest",
    2: "Grasslands",
    3: "Croplands",
    4: "Wetlands",
    5: "Artificial areas",
    6: "Other / Bare lands",
    7: "Water bodies",
}

# SOC 状态编码
# 0 = NoData / Invalid
# 1 = Degraded
# 2 = Stable
# 3 = Improved
STATUS_NODATA = 0
STATUS_DEGRADED = 1
STATUS_STABLE = 2
STATUS_IMPROVED = 3

FLOAT_NODATA = -9999.0

# 文件阈值：SOC loss/gain 10%
SOC_THRESHOLD = 0.10


# ============================================================
# 4. 面积计算函数
# ============================================================

def area_km2_from_mask(mask_array, transform, crs):
    """
    根据布尔掩膜计算面积 km²。
    支持投影坐标系，也支持 EPSG:4326 等地理坐标系下按纬度修正像元面积。
    """
    if mask_array is None or np.sum(mask_array) == 0:
        return 0.0

    res_x = transform[0]
    res_y = -transform[4]
    height = mask_array.shape[0]

    is_geographic = False
    if crs is not None:
        is_geographic = crs.is_geographic or (crs.to_epsg() == 4326)

    if is_geographic:
        R = 6371.0  # km

        y_origin = transform[5]
        pixel_height_deg = transform[4]

        row_indices = np.arange(height) + 0.5
        latitudes = y_origin + row_indices * pixel_height_deg
        lat_rad = np.radians(latitudes)

        pixel_width_km = abs(res_x) * (np.pi / 180.0) * R * np.cos(lat_rad)
        pixel_height_km = abs(res_y) * (np.pi / 180.0) * R

        row_pixel_area_km2 = pixel_width_km * pixel_height_km
        count_per_row = np.sum(mask_array, axis=1)

        return float(np.sum(count_per_row * row_pixel_area_km2))

    pixel_area_km2 = (abs(res_x) * abs(res_y)) / 1_000_000.0
    return float(np.sum(mask_array) * pixel_area_km2)


# ============================================================
# 5. 文件 Table 3.6 的完整 CF 矩阵
# ============================================================

def build_soc_cf_matrix(f_value):
    """
    完全按 GPG Addendum Table 3.6 构建 SOC land-use conversion factor matrix。

    行 = Origin class
    列 = Final class

    类别：
    1 = Forest / Tree-covered areas
    2 = Grasslands
    3 = Croplands
    4 = Wetlands
    5 = Artificial areas
    6 = Other / Bare lands
    7 = Water bodies
    """
    if f_value <= 0:
        raise ValueError("f_value 必须 > 0。")

    return {
        # Origin: Forest / Tree-covered areas
        (1, 1): 1.0,
        (1, 2): 1.0,
        (1, 3): f_value,
        (1, 4): 1.0,
        (1, 5): 0.1,
        (1, 6): 0.1,
        (1, 7): 1.0,

        # Origin: Grasslands
        (2, 1): 1.0,
        (2, 2): 1.0,
        (2, 3): f_value,
        (2, 4): 1.0,
        (2, 5): 0.1,
        (2, 6): 0.1,
        (2, 7): 1.0,

        # Origin: Croplands
        (3, 1): 1.0 / f_value,
        (3, 2): 1.0 / f_value,
        (3, 3): 1.0,
        (3, 4): 1.0 / 0.71,
        (3, 5): 0.1,
        (3, 6): 0.1,
        (3, 7): 1.0,

        # Origin: Wetlands
        (4, 1): 1.0,
        (4, 2): 1.0,
        (4, 3): 0.71,
        (4, 4): 1.0,
        (4, 5): 0.1,
        (4, 6): 0.1,
        (4, 7): 1.0,

        # Origin: Artificial areas
        (5, 1): 2.0,
        (5, 2): 2.0,
        (5, 3): 2.0,
        (5, 4): 2.0,
        (5, 5): 1.0,
        (5, 6): 1.0,
        (5, 7): 1.0,

        # Origin: Other / Bare lands
        (6, 1): 2.0,
        (6, 2): 2.0,
        (6, 3): 2.0,
        (6, 4): 2.0,
        (6, 5): 1.0,
        (6, 6): 1.0,
        (6, 7): 1.0,

        # Origin: Water bodies
        (7, 1): 1.0,
        (7, 2): 1.0,
        (7, 3): 1.0,
        (7, 4): 1.0,
        (7, 5): 1.0,
        (7, 6): 1.0,
        (7, 7): 1.0,
    }


# ============================================================
# 6. 文件路径、栅格检查与写出函数
# ============================================================

def land_cover_path(year):
    return LAND_COVER_DIR / f"{DATASET_PREFIX}_{year}_reclass_300m.tif"


def check_same_grid(src_a, src_b):
    """
    检查两个 raster 是否完全同网格。
    """
    if src_a.width != src_b.width or src_a.height != src_b.height:
        raise ValueError(
            f"栅格尺寸不一致: "
            f"A=({src_a.height}, {src_a.width}), "
            f"B=({src_b.height}, {src_b.width})"
        )

    if src_a.transform != src_b.transform:
        raise ValueError("transform 不一致，请先重采样到同一网格。")

    if str(src_a.crs) != str(src_b.crs):
        raise ValueError(f"CRS 不一致: {src_a.crs} vs {src_b.crs}")


def write_float_raster(path, array, ref_meta, nodata=FLOAT_NODATA):
    meta = ref_meta.copy()
    meta.update({
        "dtype": "float32",
        "count": 1,
        "nodata": nodata,
        "compress": "lzw",
    })

    with rasterio.open(path, "w", **meta) as dst:
        dst.write(array.astype(np.float32), 1)


def write_status_raster(path, array, ref_meta, nodata=STATUS_NODATA):
    meta = ref_meta.copy()
    meta.update({
        "dtype": "uint8",
        "count": 1,
        "nodata": nodata,
        "compress": "lzw",
    })

    with rasterio.open(path, "w", **meta) as dst:
        dst.write(array.astype(np.uint8), 1)


# ============================================================
# 7. 单个目标年份的 SOC 计算
# ============================================================

def calculate_soc_for_target_year(baseline_year, report_year):
    """
    按 GPG Addendum 的 SOC Tier 1 方法计算一个 target year。

    当前采用“两期土地覆盖图”模式：
    baseline_year land cover vs report_year land cover

    文件中说明，如果只有起始与结束两期图，则可假设变化发生在期初。
    因此：
        T = report_year - baseline_year

    核心公式：
        SOC_change_prop = ((CF - 1) * T) / 20

    判定：
        SOC_change_prop <= -0.10 -> Degraded
        -0.10 < SOC_change_prop < 0.10 -> Stable
        SOC_change_prop >= 0.10 -> Improved
    """
    origin_path = land_cover_path(baseline_year)
    final_path = land_cover_path(report_year)

    if not origin_path.exists():
        raise FileNotFoundError(f"找不到 baseline land cover 文件: {origin_path}")

    if not final_path.exists():
        raise FileNotFoundError(f"找不到 target land cover 文件: {final_path}")

    print("\n" + "=" * 90)
    print(f"开始计算 SOC: {baseline_year} -> {report_year}")
    print(f"Origin raster: {origin_path}")
    print(f"Final raster : {final_path}")

    # 按文件“两期图”假设：变化发生在期初
    T = report_year - baseline_year
    if T <= 0:
        raise ValueError("report_year 必须大于 baseline_year。")

    print(f"Climate zone = {CLIMATE_ZONE}")
    print(f"f value      = {F_VALUE}")
    print(f"T years      = {T}")

    cf_matrix = build_soc_cf_matrix(F_VALUE)

    with rasterio.open(origin_path) as src_o, rasterio.open(final_path) as src_f:
        check_same_grid(src_o, src_f)

        arr_o = src_o.read(1)
        arr_f = src_f.read(1)

        ref_meta = src_o.meta.copy()
        transform = src_o.transform
        crs = src_o.crs

    # 只有 Table 3.6 的 1-7 类参与 CF 计算
    valid_cf_mask = np.isin(arr_o, ALL_CLASSES) & np.isin(arr_f, ALL_CLASSES)

    # 用于“土地面积”统计时，通常不把 water bodies 作为有效土地面积；
    # 但水体在 CF 矩阵中仍被赋予 CF=1，不会产生 degraded/improved。
    valid_land_mask = (
        valid_cf_mask
        & np.isin(arr_o, LAND_CLASSES_FOR_AREA)
        & np.isin(arr_f, LAND_CLASSES_FOR_AREA)
    )

    # 输出 CF raster
    cf = np.full(arr_o.shape, FLOAT_NODATA, dtype=np.float32)

    for o_cls in ALL_CLASSES:
        for f_cls in ALL_CLASSES:
            mask = valid_cf_mask & (arr_o == o_cls) & (arr_f == f_cls)
            cf[mask] = cf_matrix[(o_cls, f_cls)]

    # SOC 相对变化比例
    # 例如 -0.116 表示 -11.6%
    soc_change_prop = np.full(arr_o.shape, FLOAT_NODATA, dtype=np.float32)
    soc_change_prop[valid_cf_mask] = ((cf[valid_cf_mask] - 1.0) * T) / 20.0

    # SOC 百分比变化
    # 例如 -11.6 表示 -11.6%
    soc_change_pct = np.full(arr_o.shape, FLOAT_NODATA, dtype=np.float32)
    soc_change_pct[valid_cf_mask] = soc_change_prop[valid_cf_mask] * 100.0

    # SOC status raster
    # 0 = NoData / Invalid
    # 1 = Degraded
    # 2 = Stable
    # 3 = Improved
    status = np.full(arr_o.shape, STATUS_NODATA, dtype=np.uint8)
    status[valid_cf_mask] = STATUS_STABLE

    degraded_mask = valid_cf_mask & (soc_change_prop <= -SOC_THRESHOLD)
    improved_mask = valid_cf_mask & (soc_change_prop >= SOC_THRESHOLD)
    stable_mask = valid_cf_mask & (~degraded_mask) & (~improved_mask)

    status[degraded_mask] = STATUS_DEGRADED
    status[improved_mask] = STATUS_IMPROVED
    status[stable_mask] = STATUS_STABLE

    # 输出 raster
    cf_path = OUTPUT_DIR / f"soc_cf_{baseline_year}_{report_year}.tif"
    change_prop_path = OUTPUT_DIR / f"soc_change_prop_{baseline_year}_{report_year}.tif"
    change_pct_path = OUTPUT_DIR / f"soc_change_pct_{baseline_year}_{report_year}.tif"
    status_path = OUTPUT_DIR / f"soc_status_{baseline_year}_{report_year}.tif"

    write_float_raster(cf_path, cf, ref_meta, nodata=FLOAT_NODATA)
    write_float_raster(change_prop_path, soc_change_prop, ref_meta, nodata=FLOAT_NODATA)
    write_float_raster(change_pct_path, soc_change_pct, ref_meta, nodata=FLOAT_NODATA)
    write_status_raster(status_path, status, ref_meta, nodata=STATUS_NODATA)

    # 面积统计
    degraded_area_km2 = area_km2_from_mask(degraded_mask, transform, crs)
    improved_area_km2 = area_km2_from_mask(improved_mask, transform, crs)
    stable_area_km2 = area_km2_from_mask(stable_mask & valid_land_mask, transform, crs)
    valid_land_area_km2 = area_km2_from_mask(valid_land_mask, transform, crs)

    pct_degraded = (
        degraded_area_km2 / valid_land_area_km2 * 100.0
        if valid_land_area_km2 > 0
        else 0.0
    )
    pct_improved = (
        improved_area_km2 / valid_land_area_km2 * 100.0
        if valid_land_area_km2 > 0
        else 0.0
    )
    pct_stable = (
        stable_area_km2 / valid_land_area_km2 * 100.0
        if valid_land_area_km2 > 0
        else 0.0
    )

    valid_values = soc_change_pct[valid_cf_mask]
    min_change_pct = float(np.nanmin(valid_values)) if valid_values.size > 0 else np.nan
    max_change_pct = float(np.nanmax(valid_values)) if valid_values.size > 0 else np.nan
    mean_change_pct = float(np.nanmean(valid_values)) if valid_values.size > 0 else np.nan

    # 转移组合统计表
    transition_rows = []

    for o_cls in ALL_CLASSES:
        for f_cls in ALL_CLASSES:
            mask = valid_cf_mask & (arr_o == o_cls) & (arr_f == f_cls)
            area_km2 = area_km2_from_mask(mask, transform, crs)

            if area_km2 <= 0:
                continue

            factor = cf_matrix[(o_cls, f_cls)]
            change_prop = ((factor - 1.0) * T) / 20.0
            change_pct = change_prop * 100.0

            if change_prop <= -SOC_THRESHOLD:
                soc_status = "Degraded"
            elif change_prop >= SOC_THRESHOLD:
                soc_status = "Improved"
            else:
                soc_status = "Stable"

            transition_rows.append({
                "baseline_year": baseline_year,
                "target_year": report_year,
                "origin_class_code": o_cls,
                "origin_class_name": CLASS_NAMES[o_cls],
                "final_class_code": f_cls,
                "final_class_name": CLASS_NAMES[f_cls],
                "CF": factor,
                "T_years": T,
                "SOC_change_prop": change_prop,
                "SOC_change_pct": change_pct,
                "SOC_status": soc_status,
                "area_km2": area_km2,
            })

    transition_df = pd.DataFrame(transition_rows)
    transition_csv_path = OUTPUT_DIR / f"soc_transition_summary_{baseline_year}_{report_year}.csv"
    transition_df.to_csv(transition_csv_path, index=False, encoding="utf-8-sig")

    print(f"完成: {baseline_year} -> {report_year}")
    print(f"输出 CF raster              : {cf_path}")
    print(f"输出 SOC change proportion  : {change_prop_path}")
    print(f"输出 SOC change percent     : {change_pct_path}")
    print(f"输出 SOC status raster      : {status_path}")
    print(f"输出 transition summary CSV : {transition_csv_path}")
    print(
        f"面积统计: "
        f"Degraded={degraded_area_km2:.2f} km², "
        f"Stable={stable_area_km2:.2f} km², "
        f"Improved={improved_area_km2:.2f} km², "
        f"Valid land={valid_land_area_km2:.2f} km²"
    )

    stats = {
        "Target Year": report_year,
        "Area of Degraded Land (km2)": degraded_area_km2,
        "Area of Improved Land (km2)": improved_area_km2,
        "Area of Stable Land (km2)": stable_area_km2,
        "Valid Land Area (km2)": valid_land_area_km2,
        "Percent Degraded (%)": pct_degraded,
        "Percent Stable (%)": pct_stable,
        "Percent Improved (%)": pct_improved,
        "T_years": T,
        "Climate Zone": CLIMATE_ZONE,
        "f_value": F_VALUE,
        "Min SOC Change (%)": min_change_pct,
        "Max SOC Change (%)": max_change_pct,
        "Mean SOC Change (%)": mean_change_pct,
        "CF Raster": str(cf_path),
        "SOC Change Prop Raster": str(change_prop_path),
        "SOC Change Pct Raster": str(change_pct_path),
        "SOC Status Raster": str(status_path),
        "Transition Summary CSV": str(transition_csv_path),
    }

    return stats


# ============================================================
# 8. 批量运行，并输出最终表格
# ============================================================

if __name__ == "__main__":
    all_stats = []

    print("SOC Tier 1 calculation based on GPG Addendum Table 3.6")
    print(f"Land cover folder: {LAND_COVER_DIR}")
    print(f"Dataset prefix   : {DATASET_PREFIX}")
    print(f"Baseline year    : {BASELINE_YEAR}")
    print(f"Report years     : {list(REPORT_YEARS)}")
    print(f"Climate zone     : {CLIMATE_ZONE}")
    print(f"f value          : {F_VALUE}")
    print(f"SOC threshold    : ±{SOC_THRESHOLD * 100:.0f}%")

    for report_year in REPORT_YEARS:
        stats = calculate_soc_for_target_year(
            baseline_year=BASELINE_YEAR,
            report_year=report_year,
        )
        all_stats.append(stats)

    df_stats = pd.DataFrame(all_stats)

    # 最终报告表格：与你截图一致的三列形式
    final_table = pd.DataFrame({
        "Target Year": df_stats["Target Year"],
        "Area of Degraded Land (km2)": df_stats["Area of Degraded Land (km2)"].round(2),
        "Area of Improved Land (km2)": df_stats["Area of Improved Land (km2)"].round(2),
    })

    print("\n" + "=" * 100)
    print("Table : Areal extent of degraded and improved land based on transitions in ESRI 10m")
    print("=" * 100)
    print(final_table.to_string(index=False))

    # 保存最终表格
    final_csv_path = OUTPUT_DIR / (
        f"Table_SOC_degraded_improved_ESRI10m_"
        f"{BASELINE_YEAR}_{min(REPORT_YEARS)}_{max(REPORT_YEARS)}.csv"
    )
    final_table.to_csv(final_csv_path, index=False, encoding="utf-8-sig")

    # 保存完整统计表
    full_stats_csv_path = OUTPUT_DIR / (
        f"soc_full_stats_"
        f"{BASELINE_YEAR}_{min(REPORT_YEARS)}_{max(REPORT_YEARS)}.csv"
    )
    df_stats.to_csv(full_stats_csv_path, index=False, encoding="utf-8-sig")

    print(f"\n最终表格 CSV 已保存: {final_csv_path}")
    print(f"完整统计 CSV 已保存: {full_stats_csv_path}")

    # 可选：输出 Excel
    try:
        excel_path = OUTPUT_DIR / (
            f"Table_SOC_degraded_improved_ESRI10m_"
            f"{BASELINE_YEAR}_{min(REPORT_YEARS)}_{max(REPORT_YEARS)}.xlsx"
        )

        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            final_table.to_excel(
                writer,
                sheet_name="SOC_Table",
                index=False,
                startrow=2,
            )

            ws = writer.sheets["SOC_Table"]
            ws["A1"] = "Table : Areal extent of degraded and improved land based on transitions in ESRI 10m"

            ws.column_dimensions["A"].width = 18
            ws.column_dimensions["B"].width = 34
            ws.column_dimensions["C"].width = 34

            # 简单设置表头样式
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

            header_fill = PatternFill("solid", fgColor="8CD982")
            thin_border = Border(
                left=Side(style="thin", color="BFBFBF"),
                right=Side(style="thin", color="BFBFBF"),
                top=Side(style="thin", color="BFBFBF"),
                bottom=Side(style="thin", color="BFBFBF"),
            )

            for cell in ws[3]:
                cell.fill = header_fill
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

            for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=3):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="right", vertical="center")

            ws["A1"].font = Font(bold=True, size=14)

        print(f"最终表格 Excel 已保存: {excel_path}")

    except Exception as e:
        print("\nExcel 输出失败，但 CSV 已正常保存。")
        print(f"原因: {e}")